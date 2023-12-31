# nom: William Bouchard
# date: 2023/08/31
# description: exercices semaine 2
#############################################################################################################
# Exercices 02 – Gestion d’équipe de hockey
#############################################################################################################
'''
Vous avez la charge de gérer une équipe de hockey.
Vous devez créer un programme Python pour effectuer des opérations
liées à la gestion des joueurs et des points. Assurez-vous d'appliquer les
bonnes pratiques de codage pour rendre votre code efficace et lisible.
Voici les fonctionnalités que votre programme doit inclure :
1. Ajouter un joueur :
Demandez à l'utilisateur d'entrer le nom d'un nouveau joueur et son nombre de points.
Ajoutez le nom et les points du joueur à la liste des joueurs de l'équipe.
i. Indice : Listes imbriquées
'''
import openpyxl
import pandas as pd
from numpy import select


def ajouterJoueur(liste, wb):
    nom = input("Entrez le nom du joueur: ")
    points = int(input("Entrez le nombre de points du joueur: "))
    wb.active.append([nom, points])
    wb.save("Exercice 02.03 - Canadiens de Montreal.xlsx")
    return liste


'''
Afficher les joueurs :
Affichez la liste complète des joueurs de l'équipe.
Chaque ligne contient le nom du joueur et son nombre de points.
'''


def afficherListe(liste):
    Df2 = liste.sort_values(by=['Points'], ascending=False).dropna(axis=0)
    print(Df2)


'''
Modifier le nom d’un joueur :
Demandez à l'utilisateur d'entrer le nom d'un joueur à modifier.
Demandez à l'utilisateur d'entrer le nouveau nom du joueur.
Si le nom du joueur est présent dans la liste, mettez-le à jour.
'''


def modifierNom(liste, wb):
    nom = input("Entrez le nom du joueur à modifier: ")
    if liste['Nom du joueur'].isin([nom]).any():
        nouveauNom = input("Entrez le nouveau nom du joueur: ")
        for row in wb.active.iter_rows():
            for cell in row:
                if cell.value == nom:
                    cell.value = nouveauNom
    else:
        print("Le joueur n'existe pas")

    wb.save("Exercice 02.03 - Canadiens de Montreal.xlsx")
    return liste


'''
Modifier les points :
Demandez à l'utilisateur d'entrer le nom d'un joueur pour modifier ses points.
Demandez à l'utilisateur d'entrée les nouveaux points du joueur.
Si le nom du joueur est présent dans la liste, mettez-les points à jour.
'''


def ModifierPoints(liste):
    nom = input("Entrez le nom du joueur à modifier: ")
    if liste['Nom du joueur'].isin([nom]).any():

        points = int(input("Entrez le nouveau nombre de points du joueur: "))

        for row in wb.active.iter_rows():
            for cell in row:
                if cell.value == nom:
                    wb["Feuil1"].cell(row=cell.row, column=2).value = points
            else:
                print("Le joueur n'existe pas")
    else:
        print("Le joueur n'existe pas")

    # sauvegarder le fichier
    wb.save("Exercice 02.03 - Canadiens de Montreal.xlsx")
    return liste


'''
Statistiques de l’équipe :
Affichez le nom du meilleur joueur et ses points.
i. Le nombre de joueurs dans l’équipe.
ii. Trouvez le joueur avec les points le plus élevé dans la liste.
iii. Trouvez le joueur avec les points le plus bas dans la liste.
iv. Trouvez la moyenne de points des joueurs.
'''


def Statistiques(liste):
    print("Le nombre de joueurs dans l'équipe est de: ", len(liste))
    if len(liste) >= 1:
        df2 = liste['Nom du joueur'].values[liste['Points'].values.argmin()]
        print("Le joueur avec le plus de points est: ",
              liste['Nom du joueur'].values[liste['Points'].values.argmax()])
        print("Le joueur avec le moins de points est: ",
              df2)
        print("La moyenne de points des joueurs est de: ", liste['Points'].mean())


'''
Supprimer un joueur :
Demandez à l'utilisateur d'entrer le nom d'un joueur à supprimer.
Si le joueur est présent dans la liste, supprimez-le.
'''


def supprimerJoueur(liste, wb, feuille):
    nom = input("Entrez le nom du joueur à supprimer: ")
    for row in wb.active.iter_rows():
        for cell in row:
            if cell.value == nom:
                feuille.cell(row=cell.row, column=1).value = ""
                feuille.cell(row=cell.row, column=2).value = ""
            else:
                print("Le joueur n'existe pas")
    # sauvegarder le fichier
    wb.save("Exercice 02.03 - Canadiens de Montreal.xlsx")
    return liste


def AffichageJoueur():
    choixBase = -1
    print("\nGestionnaire équipe de hockey")
    print("-----------------------------")
    print("1. Ajouter un joueur ")
    print("2. Afficher les Joueurs")
    print("3. Modifier le nom d’un joueur ")
    print("4. Modifier les points ")
    print("5. Statistiques de l’équipe ")
    print("6. Supprimer un joueur ")
    print("7. Quitter")
    choixBase = input("\nChoisissez une option :")
    return choixBase


if __name__ == "__main__":
    choixOption = -1

    while choixOption != "7":
        liste = pd.read_excel("Exercice 02.03 - Canadiens de Montreal.xlsx")
        wb = openpyxl.load_workbook("Exercice 02.03 - Canadiens de Montreal.xlsx")
        feuille = wb["Feuil1"]
        choixOption = AffichageJoueur()
        match choixOption:
            case "1":
                ajouterJoueur(liste, wb)

            case "2":
                afficherListe(liste)

            case "3":
                modifierNom(liste, wb)

            case "4":
                ModifierPoints(liste)

            case "5":
                Statistiques(liste)

            case "6":
                supprimerJoueur(liste, wb, feuille)

            case "7":
                choix = "7"
                print("Au revoir!")